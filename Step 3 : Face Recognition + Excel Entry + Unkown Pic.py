import cv2
from datetime import datetime
import os

# Create the "Unknown Faces" folder if it doesn't exist
unknown_faces_folder = "Unknown Faces"
if not os.path.exists(unknown_faces_folder):
    os.makedirs(unknown_faces_folder)


def draw_boundary(img, classifier, scaleFactor, minNeighbor, color, text, clf):
    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    features = classifier.detectMultiScale(gray_img, scaleFactor, minNeighbor)
    coords = []
    for (x, y, w, h) in features:
        cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
        id, pred = clf.predict(gray_img[y:y + h, x:x + w])
        confidence = int(100 * (1 - pred / 300))

        if confidence > 77:
            if id in users_dict:  # Check if the detected ID is in the registered users dictionary
                name = users_dict[id]

                cv2.putText(img, f"{name} {confidence}%", (x, y - 4), cv2.FONT_HERSHEY_COMPLEX, 0.8, color, 1,
                            cv2.LINE_AA)
        else:
            # Capture and save image for unknown person
            save_unknown_person(img[y:y + h, x:x + w])
            
            name = "UNKNOWN"
            cv2.putText(img, f"{name} {confidence}%", (x, y - 5), cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 0, 255), 1,
                        cv2.LINE_AA)

        coords = [x, y, w, h]

    return coords


def save_unknown_person(img):
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"unknown_{current_time}.jpg"
    file_path = os.path.join(unknown_faces_folder, file_name)
    cv2.imwrite(file_path, img)
    print(f"Saved image: {file_path}")


from openpyxl import Workbook, load_workbook


# Function to update Excel sheet with detected user's name and time
def update_excel_sheet(name, time):
    wb = Workbook()
    ws = wb.active

    # Check if the Excel file already exists, if not, create a new one
    try:
        wb = load_workbook("user_activity.xlsx")
        ws = wb.active
    except FileNotFoundError:
        pass

    # Append data to the Excel sheet
    ws.append([name, time])
    wb.save("user_activity.xlsx")


def recognize(img, clf, faceCascade, last_update_time, register_user):
    color = {"blue": (255, 0, 0), "red": (0, 0, 255), "green": (0, 255, 0), "White": (255, 255, 255)}

    coords = draw_boundary(img, faceCascade, 1.1, 10, color['White'], "Face", clf)

    if coords:
        x, y, w, h = coords
        gray_img = cv2.cvtColor(img[y:y + h, x:x + w], cv2.COLOR_BGR2GRAY)
        id, pred = clf.predict(gray_img)
        confidence = int(100 * (1 - pred / 300))

        if confidence > 77:
            if id in users_dict:  # Check if the detected ID is in the registered users dictionary
                name = users_dict[id]  # Get the user name corresponding to the ID
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if time.time() - last_update_time >= 2:
                    update_excel_sheet(name, current_time)
                    last_update_time = time.time()
                cv2.putText(img, f"{name} {confidence}%", (x, y - 4), cv2.FONT_HERSHEY_COMPLEX, 0.8, color['White'],
                            1, cv2.LINE_AA)
            else:
                name = "UNKNOWN"
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if time.time() - last_update_time >= 2:
                    update_excel_sheet(name, current_time)
                    last_update_time = time.time()
                cv2.putText(img, f"{name} {confidence}%", (x, y - 5), cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 0, 255), 1,
                            cv2.LINE_AA)

    return img, last_update_time


def detect(img, faceCascade):
    color = {"blue": (255, 0, 0), "red": (0, 0, 255), "green": (0, 255, 0), "White": (255, 255, 255)}
    coords, img = draw_boundary(img, faceCascade, 1.5, 14, color['blue'], "Face")

    if len(coords) == 4:
        roi_img = img[coords[1]:coords[1] + coords[3], coords[0]:coords[0] + coords[2]]
        # OR  -->  roi_img = img[y:y+4,x:x+4]

        user_id = 1
        generate_dataset(roi_img, user_id, img_id)

    return img


faceCascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
clf = cv2.face.LBPHFaceRecognizer_create()
clf.read("classifier1.yml")

video_capture = cv2.VideoCapture(0)

img_id = 0
last_update_time = time.time()  # Initialize last_update_time
while True:
    ret, img = video_capture.read()
    if ret:  # Check if frame is captured successfully
        img, last_update_time = recognize(img, clf, faceCascade, last_update_time, users_dict)  # Pass last_update_time
        cv2.imshow("Face Recognition", img)
        img_id += 1
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()
